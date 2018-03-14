from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment,Font,colors
from personalia.models import Member
from datetime import datetime as dt
import psycopg2

def db_server():
    try:
        conn = psycopg2.connect("dbname='barokah' user='sodikin' host='localhost' password='nurs4d4ni'")
        return conn
    except:
        print("I am unable to connect to the database")

def ConstructReport(response,params):
    report_name = params.nama_laporan
    file = open('enterprise/template_xlsx/member_detail.xlsx', 'rb')

    wb = load_workbook(filename = file)
    ws = wb['data']
    ws_pm = wb['peta member']

    #set header data
    rows = 0
    rows += 1;ws.cell(row=rows,column=1,value=report_name.upper())
    date_now = dt.today()
    date_now = date_now.strftime("%d %B %Y")
    rows += 1;ws.cell(row=rows,column=1,value=date_now)
    rows += 2

    member = Member.objects.all().order_by('jenjang','individu__jk','pangkal','individu__nama')

    for m in member:
        col = 0
        rows += 1
        ft = Font(color=colors.RED)
        ac = Alignment(horizontal='center')
        col += 1;ws.cell(row=rows,column=col,value=rows-4).alignment = ac
        col += 1;ws.cell(row=rows,column=col,value=m.individu.nama.upper())
        col += 1;ws.cell(row=rows,column=col,value=m.individu.get_jk_display())
        col += 1;ws.cell(row=rows,column=col,value=m.individu.tgl_lahir).number_format = 'dd mmm yyy'
        col += 1;ws.cell(row=rows,column=col,value=m.get_jenjang_display()).alignment = ac
        col += 1;ws.cell(row=rows,column=col,value=m.get_status_aktif_display()).alignment = ac
        col += 1;ws.cell(row=rows,column=col,value=m.pangkal.nama.upper())

    #set header peta member
    rows = 0
    rows += 1;ws_pm.cell(row=rows,column=1,value='PETA MEMBER')
    date_now = dt.today()
    date_now = date_now.strftime("%d %B %Y")
    rows += 1;ws_pm.cell(row=rows,column=1,value=date_now)
    rows = 6

    conn = db_server()
    cur = conn.cursor()
    #member aktif
    cur.execute('''
        select p.nama,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '0' then '1'
        	end) as wb_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '0' then '1'
        	end) as wb_p,
        count(
        	case
        		when upper(m.jenjang) = '0' then '1'
        	end) as wb_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '1' then '1'
        	end) as pra_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '1' then '1'
        	end) as pra_p,
        count(
        	case
        		when upper(m.jenjang) = '1' then '1'
        	end) as pra_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '2' then '1'
        	end) as a1_1_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '2' then '1'
        	end) as a1_1_p,
        count(
        	case
        		when upper(m.jenjang) = '2' then '1'
        	end) as a1_1_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '3' then '1'
        	end) as a1_2_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '3' then '1'
        	end) as a1_2_p,
        count(
        	case
        		when upper(m.jenjang) = '3' then '1'
        	end) as a1_2_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '4' then '1'
        	end) as a1_3_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '4' then '1'
        	end) as a1_3_p,
        count(
        	case
        		when upper(m.jenjang) = '4' then '1'
        	end) as a1_3_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '5' then '1'
        	end) as a2_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '5' then '1'
        	end) as a2_p,
        count(
        	case
        		when upper(m.jenjang) = '5' then '1'
        	end) as a2_t
        from barokah.personalia_member m
        inner join barokah.personalia_individu i on i.id = m.individu_id
        inner join barokah.personalia_pangkal p on p.id_pk = m.pangkal_id
        where m.status_aktif = 'ak'
        group by p.nama
        order by p.nama
    ''')
    member = cur.fetchall()
    for m in member:
        col = 0
        rows += 1
        ft = Font(color=colors.RED)
        ac = Alignment(horizontal='center')
        col += 1;ws_pm.cell(row=rows,column=col,value=rows-6).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[0])
        col += 1;ws_pm.cell(row=rows,column=col,value=m[1]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[2]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[3]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[4]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[5]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[6]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[7]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[8]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[9]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[10]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[11]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[12]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[13]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[14]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[15]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[16]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[17]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[18]).alignment = ac

    #member pasif
    rows = 15
    cur2 = conn.cursor()
    cur2.execute('''
        select p.nama,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '0' then '1'
        	end) as wb_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '0' then '1'
        	end) as wb_p,
        count(
        	case
        		when upper(m.jenjang) = '0' then '1'
        	end) as wb_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '1' then '1'
        	end) as pra_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '1' then '1'
        	end) as pra_p,
        count(
        	case
        		when upper(m.jenjang) = '1' then '1'
        	end) as pra_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '2' then '1'
        	end) as a1_1_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '2' then '1'
        	end) as a1_1_p,
        count(
        	case
        		when upper(m.jenjang) = '2' then '1'
        	end) as a1_1_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '3' then '1'
        	end) as a1_2_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '3' then '1'
        	end) as a1_2_p,
        count(
        	case
        		when upper(m.jenjang) = '3' then '1'
        	end) as a1_2_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '4' then '1'
        	end) as a1_3_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '4' then '1'
        	end) as a1_3_p,
        count(
        	case
        		when upper(m.jenjang) = '4' then '1'
        	end) as a1_3_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '5' then '1'
        	end) as a2_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '5' then '1'
        	end) as a2_p,
        count(
        	case
        		when upper(m.jenjang) = '5' then '1'
        	end) as a2_t
        from barokah.personalia_member m
        inner join barokah.personalia_individu i on i.id = m.individu_id
        inner join barokah.personalia_pangkal p on p.id_pk = m.pangkal_id
        where m.status_aktif = 'pa'
        group by p.nama
        order by p.nama
    ''')
    member = cur2.fetchall()
    for m in member:
        col = 0
        rows += 1
        ft = Font(color=colors.RED)
        ac = Alignment(horizontal='center')
        col += 1;ws_pm.cell(row=rows,column=col,value=rows-15).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[0])
        col += 1;ws_pm.cell(row=rows,column=col,value=m[1]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[2]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[3]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[4]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[5]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[6]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[7]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[8]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[9]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[10]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[11]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[12]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[13]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[14]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[15]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[16]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[17]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[18]).alignment = ac

    #member bk
    rows = 24
    cur3 = conn.cursor()
    cur3.execute('''
        select p.nama,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '0' then '1'
        	end) as wb_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '0' then '1'
        	end) as wb_p,
        count(
        	case
        		when upper(m.jenjang) = '0' then '1'
        	end) as wb_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '1' then '1'
        	end) as pra_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '1' then '1'
        	end) as pra_p,
        count(
        	case
        		when upper(m.jenjang) = '1' then '1'
        	end) as pra_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '2' then '1'
        	end) as a1_1_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '2' then '1'
        	end) as a1_1_p,
        count(
        	case
        		when upper(m.jenjang) = '2' then '1'
        	end) as a1_1_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '3' then '1'
        	end) as a1_2_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '3' then '1'
        	end) as a1_2_p,
        count(
        	case
        		when upper(m.jenjang) = '3' then '1'
        	end) as a1_2_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '4' then '1'
        	end) as a1_3_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '4' then '1'
        	end) as a1_3_p,
        count(
        	case
        		when upper(m.jenjang) = '4' then '1'
        	end) as a1_3_t,
        count(
        	case
        		when upper(i.jk) = 'L' and upper(m.jenjang) = '5' then '1'
        	end) as a2_l,
        count(
        	case
        		when upper(i.jk) = 'P' and upper(m.jenjang) = '5' then '1'
        	end) as a2_p,
        count(
        	case
        		when upper(m.jenjang) = '5' then '1'
        	end) as a2_t
        from barokah.personalia_member m
        inner join barokah.personalia_individu i on i.id = m.individu_id
        inner join barokah.personalia_pangkal p on p.id_pk = m.pangkal_id
        where m.status_aktif in ('bk1','bk2','bk3')
        group by p.nama
        order by p.nama
    ''')
    member = cur3.fetchall()
    for m in member:
        col = 0
        rows += 1
        ft = Font(color=colors.RED)
        ac = Alignment(horizontal='center')
        col += 1;ws_pm.cell(row=rows,column=col,value=rows-24).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[0])
        col += 1;ws_pm.cell(row=rows,column=col,value=m[1]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[2]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[3]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[4]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[5]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[6]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[7]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[8]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[9]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[10]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[11]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[12]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[13]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[14]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[15]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[16]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[17]).alignment = ac
        col += 1;ws_pm.cell(row=rows,column=col,value=m[18]).alignment = ac

    conn.close()

    wb.save(response)
    return response
